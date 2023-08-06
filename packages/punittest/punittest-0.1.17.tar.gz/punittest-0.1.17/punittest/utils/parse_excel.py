#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

from .logger import logger
from openpyxl.styles import PatternFill


def create_excel_file(file_path):
    """
    创建xlsx表格
    :param file_path: 表格全路径
    :return: None
    """
    wb = openpyxl.Workbook()
    wb.create_sheet(title="Sheet1", index=0)
    wb.save(file_path)


class ExcelParser:

    def __init__(self, excel_file_path, sheet_name):
        self.excel_file_path = excel_file_path
        self.workbook = openpyxl.load_workbook(self.excel_file_path)
        self.worksheet = self.workbook[sheet_name]
        self.max_row = self.worksheet.max_row
        self.max_column = self.worksheet.max_column

    def save_workbook(self):
        self.workbook.save(self.excel_file_path)

    def create_new_sheet(self, new_sheet_name="new_sheet"):
        try:
            self.workbook.create_sheet(new_sheet_name)
            self.workbook.save(self.excel_file_path)
        except Exception as e:
            logger.error("Fail to create new sheet: {0}".format(e)[:-1])
            raise e

    def get_cell_value(self, row, column):
        return self.worksheet.cell(row=row, column=column).value

    def set_cell_value(self, row, column, value):
        self.worksheet.cell(row=row, column=column).value = value
        return value

    def get_row_values(self, row, filter_none='end'):
        row_values = [self.get_cell_value(row, col) for col in range(1, self.max_column + 1)]
        if filter_none in ['all', 'All', 'ALL']:
            row_values = list(filter(lambda value: value is not None, row_values))
        elif filter_none in ['end', 'End', 'END']:
            for i, val in enumerate(row_values[::-1]):
                if val is not None:
                    row_values = row_values[:len(row_values)-i]
                    break
        return row_values

    def get_col_values(self, col, filter_none='end'):
        col_values = [self.get_cell_value(row, col) for row in range(1, self.max_row + 1)]
        if filter_none in ['all', 'All', 'ALL']:
            col_values = list(filter(lambda value: value is not None, col_values))
        elif filter_none in ['end', 'End', 'END']:
            for i, val in enumerate(col_values[::-1]):
                if val is not None:
                    col_values = col_values[:len(col_values)-i]
                    break
        return col_values

    def set_cell_background_color(self, row, column, fgColor):
        fill = PatternFill(fgColor=fgColor, fill_type='solid')
        self.worksheet.cell(row=row, column=column).fill = fill
        return True


if __name__ == '__main__':
    pass
