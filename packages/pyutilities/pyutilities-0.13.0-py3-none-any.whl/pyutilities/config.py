#!/usr/bin/env python
# coding=utf-8

"""
    Utility class for holding configuration. Can merge configuration with environment variables.
    Can load configuration from YAML files. See docstring for Configuration class.

    18.11.2018
    Added child config class that is able to load config (as dictionary) from xls file (from specified sheet).

    Created:  Gusev Dmitrii, XX.08.2017
    Modified: Gusev Dmitrii, 10.01.2021

"""

import os
import logging
import xlrd  # reading excel files (old Excel - xls)
import xlwt  # writing excel files
import openpyxl  # reading excel files (Excel 2010 - xlsx)

from string import Template
from pyutilities.utils import parse_yaml

YAML_EXTENSION_1 = '.yml'
YAML_EXTENSION_2 = '.yaml'
DEFAULT_ENCODING = 'UTF8'


class Configuration(object):
    """Tree-like configuration-holding structure, allows loading from YAML and retrieving values
    by using chained hierarchical key with dot-separated levels, e.g. "hdfs.namenode.address".
    Can include environment variables (switch by key), environment usually override internal values.
    :param path_to_config ???
    :param dict_to_merge ???
    :param is_override_config ???
    :param is_merge_env ???
    """

    def __init__(self, path_to_config=None, dict_to_merge=None, is_override_config=True, is_merge_env=True):
        # init logger
        self.log = logging.getLogger(__name__)
        self.log.addHandler(logging.NullHandler())
        self.log.debug("Initializing Configuration() instance...")
        self.log.debug("Load configuration:\n\tpath -> {}\n\tdict -> {}\n\toverride config -> {}\n\tmerge env -> {}"
                       .format(path_to_config, dict_to_merge, is_override_config, is_merge_env))

        # init internal dictionary
        self.config_dict = {}

        if path_to_config and path_to_config.strip():  # if provided file path - try to load config
            self.log.debug("Loading config from [{}].".format(path_to_config))
            self.load(path_to_config, is_merge_env)

        if dict_to_merge:  # merge config from file(s) (if any) with dictionary, if any
            self.log.debug("Dictionary for merge isn't empty: {}".format(dict_to_merge))
            if isinstance(dict_to_merge, dict):  # provided single dictionary
                self.log.debug("Merging with provided single dictionary. Override: [{}].".format(is_override_config))
                self.append_dict(dict_to_merge, is_override_config)
            elif isinstance(dict_to_merge, list):  # provided list of dictionaries
                self.log.debug("Merging with provided list of dictionaries. Override: [{}].".format(is_override_config))
                for dictionary in dict_to_merge:
                    self.append_dict(dictionary, is_override_config)
            else:
                raise ConfigError('Provided unknown type [{}] of dictionary for merge!'.format(type(dict_to_merge)))
        self.log.info("Configuration loaded successfully.")

    def append_dict(self, dictionary, is_override):
        """Appends specified dictionary to internal dictionary of current class.
            :param dictionary
            :param is_override
        """
        self.log.debug("append_dict() is working.")
        for key, value in dictionary.items():
            if is_override or key not in self.config_dict.keys():
                # override key only with non-empty value
                if value:
                    self.set(key, value)

    def load(self, path, is_merge_env=True):
        """Parses YAML file(s) from the given directory/file to add content into this configuration instance
            :param is_merge_env: merge parameters with environment (True) or not (False)
            :param path: directory/file to load files from
            :type path: str
        """
        self.log.debug('load() is working. Path [{}], is_merge_env [{}].'.format(path, is_merge_env))
        # fail-fast checks
        if not path or not path.strip():
            raise ConfigError('Provided empty path for config loading!')
        if not os.path.exists(path):
            raise ConfigError('Provided path [%s] doesn\'t exist!' % path)

        # todo: extract two methods - load from file/load from dir + refactor unit tests
        # if provided path to single file - load it, otherwise - load from directory
        if os.path.isfile(path) and (path.endswith(YAML_EXTENSION_1) or path.endswith(YAML_EXTENSION_2)):
            self.log.debug("Provided path [{}] is a YAML file. Loading.".format(path))
            try:
                self.merge_dict(parse_yaml(path))
            except ConfigError as ex:
                raise ConfigError('ERROR while merging file %s to configuration.\n%s' % (path, ex.message))

        # loading from directory (all YAML files)
        elif os.path.isdir(path):
            self.log.debug("Provided path [{}] is a directory. Loading all YAML files.".format(path))
            for some_file in os.listdir(path):
                file_path = os.path.join(path, some_file)
                if os.path.isfile(file_path) and \
                        (some_file.endswith(YAML_EXTENSION_1) or some_file.endswith(YAML_EXTENSION_2)):
                    self.log.debug("Loading configuration from [{}].".format(some_file))
                    try:
                        self.merge_dict(parse_yaml(file_path))
                    except ConfigError as ex:
                        raise ConfigError('ERROR while merging file %s to configuration.\n%s' % (file_path, ex.message))

        # unknown file/dir type
        else:
            raise ConfigError('Unknown thing [%s], not a file, not a dir!' % path)

        # merge environment variables to internal dictionary
        if is_merge_env:
            self.log.info("Merging environment variables is switched ON.")
            self.merge_env()

    # todo: possible bug: if merge dict with multi-level keys (a.b.c), these keys can't be accessed by get() method!
    def merge_dict(self, new_dict):
        """Adds another dictionary (respecting nested sub-dictionaries) to config.
        If there are same keys in both dictionaries, raise ConfigError (no overwrites!)
        :param new_dict: dictionary to be added
        :type new_dict: dict
        """
        self.log.debug("merge_dict() is working. Dictionary to merge [{}].".format(new_dict))
        dict1 = self.config_dict
        if len(dict1) != 0:
            result = self.__add_entity__(dict1, new_dict)
            self.config_dict = result
        else:
            self.config_dict.update(new_dict)
        return

    def __add_entity__(self, dict1, dict2, current_key=''):
        """Adds second dictionary to the first (processing nested dicts recursively)
        No overwriting is accepted.
        :param dict1: target dictionary (exception raising if it is not dict)
        :type dict1: dict
        :param dict2: source dictionary (exception raising if it is not dict)
        :type dict2: dict
        :return: the first parameter (i.e. target dictionary)
        :rtype: dict
        """
        if isinstance(dict2, dict) and isinstance(dict1, dict):
            for key in dict2.keys():
                if key in dict1.keys():
                    sep = '.'
                    if current_key == '':
                        sep = ''
                    self.__add_entity__(dict1[key], dict2[key], '%s%s%s' % (current_key, sep, key))
                else:
                    dict1[key] = dict2[key]
        else:
            raise ConfigError("Attempt of overwriting old value %s with new %s to the key %s" % (dict1, dict2, current_key))
        return dict1

    def merge_env(self):
        """Adds environment variables to this config instance"""
        self.log.debug("merge_env() is working.")
        for item in os.environ:
            self.config_dict[item.lower()] = os.environ[item]

    def get(self, key, default=None):
        """Retrieves config value for given key.
            :param key: text key, which could be complex, e.g. "key1.key2.key3" to reach nested dictionaries
            :type key: str
            :type default: Any
            :rtype: Any
        """
        try:
            result = self.__get_value(key, self.config_dict)
            return result
        except KeyError as err:
            if default is not None:
                return default
            raise ConfigError("Configuration entry [{}] not found".format(key))

    def set(self, key, value):
        """Sets config value, creating all the nested levels if necessary
            :type key: str
            :type value: Any
        """
        keys = key.split(".")
        values = self.config_dict
        while len(keys) > 1:
            cur = keys.pop(0)
            if cur not in values:
                values[cur] = dict()
            values = values[cur]
        values[keys[0]] = value

    def resolve_and_set(self, key, value):
        """Performs template substitution in "value" using mapping from config (only top-level), then sets it in config
            :param key: key to assign value to (could be multi-level)
            :type key: str
            :param value: value with substitution patterns e.g. "system-$env" (see string.Template)
            :type value: str 
        """
        template = Template(value)
        resolved = template.substitute(self.config_dict)
        self.set(key, resolved)

    def __get_value(self, key, values):
        """Internal method for using in "get", recursively search for dictionaries by key parts and retrieves value
            :type key: str
            :param values: dictionary to search in
            :type values: dict
            :rtype: Any
        """
        if not values:
            raise KeyError
        keys = key.split(".", 1)
        if len(keys) < 2:
            return values[keys[0]]
        else:
            return self.__get_value(keys[1], values[keys[0]])

    def __contains_key(self, key, values):
        if not values:
            return False
        keys = key.split(".", 1)
        if len(keys) < 2:
            return keys[0] in values
        else:
            return self.__contains_key(keys[1], values[keys[0]])

    def contains_key(self, key):
        return self.__contains_key(key, self.config_dict)

    def __str__(self):
        return str(self.config_dict)


class ConfigError(Exception):
    """Invalid configuration error"""


# numbers of name/value columns in excel config sheet
# NOTE! For excel < 2010 (xls) - zero-based, for excel >= 2010 (xlsx) - one-based.
#       For excel < 2010 (xls) used xlrd lib, for excel >= 2010 (xlsx) used openpyxl lib.
NAMES_COLUMN = 0
VALUES_COLUMN = 1


class ConfigurationXls(Configuration):
    """Extension for Configuration class for work with excel config files.
    """

    def __init__(self, path_to_xls, config_sheet_name, dict_to_merge=None, path_to_yaml=None,
                 is_override_config=True, is_merge_env=True):
        # init class instance logger
        self.log = logging.getLogger(__name__)
        self.log.addHandler(logging.NullHandler())
        self.log.debug("Initializing ConfigurationXls() instance...")  # <- INFO level changed to DEBUG

        # load config (dictionary) from xls file
        xls_dict = self.load_dict_from_xls(path_to_xls, config_sheet_name)

        # calculate dictionaries for merge
        dictionary = []
        if dict_to_merge is None:  # dictionary for merge doesn't provided
            dictionary = xls_dict
        elif isinstance(dict_to_merge, dict):  # provided dictionary for merge is dictionary
            dictionary.append(dict_to_merge)
            dictionary.append(xls_dict)
        elif isinstance(dict_to_merge, list):  # provided dictionary for merge is a list of dictionaries
            dictionary = dict_to_merge
            dictionary.append(xls_dict)
        else:  # provided invalid type of dictionary for merge
            raise ConfigError("Invalid dictionary to merge type: {}!".format(type(dict_to_merge)))

        # init instance with superclass constructor
        super(ConfigurationXls, self).__init__(path_to_config=path_to_yaml, dict_to_merge=dictionary,
                                               is_override_config=is_override_config, is_merge_env=is_merge_env)

    def load_dict_from_xls(self, path_to_xls, config_sheet_name):
        self.log.debug("load_dict_from_xls() is working.")
        self.log.debug("Excel file [{}], config sheet [{}].".format(path_to_xls, config_sheet_name))

        # some preliminary checks (fast-fail)
        if not path_to_xls or not path_to_xls.strip():
            raise ConfigError('Provided empty path to xls file!')
        if not os.path.exists(path_to_xls):
            raise ConfigError('Provided path [%s] doesn\'t exist!' % path_to_xls)

        dictionary = {}  # dictionary loaded from excel file

        # loading xls/xlsx workbook
        if path_to_xls.endswith('xls'):  # load from excel file - format xls
            excel_book = xlrd.open_workbook(path_to_xls, encoding_override=DEFAULT_ENCODING)
            excel_sheet = excel_book.sheet_by_name(config_sheet_name)
            self.log.debug("Loaded XLS config. Found [{}] row(s). Loading.".format(excel_sheet.nrows))
            # loading dictionary from xls file
            for rownumber in range(excel_sheet.nrows):
                name = excel_sheet.cell_value(rownumber, NAMES_COLUMN)
                value = excel_sheet.cell_value(rownumber, VALUES_COLUMN)
                self.log.debug("XLS: loaded config parameter: {} = {}".format(name, value))
                dictionary[name] = value

        elif path_to_xls.endswith('xlsx'):  # load from excel file - format xlsx
            excel_book = openpyxl.load_workbook(path_to_xls)  # workbook
            excel_sheet = excel_book[config_sheet_name]  # specified sheet by name
            self.log.debug("Loaded XSLX config. Found [{}] row(s). Loading.".format(excel_sheet.max_row))
            for rownumber in range(excel_sheet.max_row):
                name = excel_sheet.cell(row=rownumber + 1, column=NAMES_COLUMN + 1).value
                value = excel_sheet.cell(row=rownumber + 1, column=VALUES_COLUMN + 1).value
                self.log.debug("XLSX: loaded config parameter: {} = {}".format(name, value))
                dictionary[name] = value

        else:  # unknown extension of excel file - raise an issue
            raise ConfigError("Provided unknown excel file extension [%s]!" % path_to_xls)

        self.log.info("Loaded dictionary from xls config:\n\t{}".format(dictionary))
        return dictionary


if __name__ == '__main__':
    print("pyutilities.config: Don't try to execute library as a standalone app!")
